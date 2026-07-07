"""Build a local SQLite database from the public MycoBank MBList export."""

from __future__ import annotations

import hashlib
import re
import shutil
import sqlite3
import sys
import tempfile
import unicodedata
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Iterable
from urllib.request import Request, urlopen
from zipfile import ZipFile

from openpyxl import load_workbook

from .contracts import (
    DEFAULT_CACHE_DIR_NAME,
    DEFAULT_EXCEL_FILE,
    DEFAULT_MYCOBANK_URL,
    DEFAULT_SHEET,
    DEFAULT_TABLE,
    DEFAULT_USER_AGENT,
)


def quote_ident(identifier: str) -> str:
    """Quote SQLite identifiers safely."""
    return '"' + identifier.replace('"', '""') + '"'


def clean_column_name(value: object, position: int, used: set[str]) -> str:
    """Convert Excel headers into stable SQLite column names."""
    raw = "" if value is None else str(value).strip()

    if not raw:
        name = f"column_{position}"
    else:
        name = unicodedata.normalize("NFKD", raw)
        name = name.encode("ascii", "ignore").decode("ascii")
        name = name.lower()
        name = re.sub(r"[^a-z0-9_]+", "_", name)
        name = re.sub(r"_+", "_", name).strip("_")

    if not name:
        name = f"column_{position}"
    if name[0].isdigit():
        name = f"col_{name}"

    base = name
    suffix = 2
    while name in used:
        name = f"{base}_{suffix}"
        suffix += 1
    used.add(name)
    return name


def sqlite_value(value: object) -> object:
    """Normalize Excel cell values before insertion."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def download_file(url: str, destination: Path, *, refresh: bool = False) -> Path:
    """Download ``url`` to ``destination`` unless it already exists."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists() and not refresh:
        return destination

    request = Request(url, headers={"User-Agent": DEFAULT_USER_AGENT})
    with urlopen(request, timeout=120) as response:
        with destination.open("wb") as fh:
            shutil.copyfileobj(response, fh)
    return destination


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        while chunk := fh.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def resolve_zip_path(
    *,
    zip_path: Path | None,
    url: str,
    cache_dir: Path,
    refresh: bool,
) -> Path:
    """Return a local MBList ZIP path from a manual file or remote URL."""
    if zip_path is not None:
        if not zip_path.exists():
            raise FileNotFoundError(f"No existe el ZIP local: {zip_path}")
        return zip_path

    return download_file(url, cache_dir / "MBList.zip", refresh=refresh)


def extract_mblist_excel(zip_path: Path, destination_dir: Path) -> tuple[Path, str]:
    """Extract the expected Excel workbook from an MBList ZIP."""
    destination_dir.mkdir(parents=True, exist_ok=True)
    with ZipFile(zip_path) as zf:
        excel_members = [
            member
            for member in zf.namelist()
            if member.lower().endswith((".xlsx", ".xlsm", ".xls"))
            and not Path(member).name.startswith("~$")
            and "__macosx" not in member.lower()
        ]
        if not excel_members:
            raise ValueError(f"No se encontró ningún Excel dentro de {zip_path}")

        exact_matches = [
            member for member in excel_members if Path(member).name == DEFAULT_EXCEL_FILE
        ]
        if exact_matches:
            member = exact_matches[0]
        elif len(excel_members) == 1:
            member = excel_members[0]
        else:
            raise ValueError(
                "El ZIP contiene varios Excel y no se puede decidir cuál cargar: "
                + ", ".join(excel_members)
            )

        output_path = destination_dir / Path(member).name
        with zf.open(member) as source, output_path.open("wb") as target:
            shutil.copyfileobj(source, target)
    return output_path, member


def worksheet_has_data(ws) -> bool:
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            return True
    return False


def get_header_row(rows: Iterable[tuple[object, ...]]) -> tuple[object, ...]:
    for row in rows:
        if any(cell is not None for cell in row):
            return row
    raise ValueError("La hoja no contiene ninguna fila con cabeceras.")


def prepare_database(
    conn: sqlite3.Connection,
    *,
    table_name: str,
    columns: list[str],
    original_headers: list[str],
    replace: bool,
) -> None:
    cur = conn.cursor()
    if replace:
        cur.execute(f"DROP TABLE IF EXISTS {quote_ident(table_name)}")
        cur.execute("DROP TABLE IF EXISTS _mycobank_import_meta")
        cur.execute("DROP TABLE IF EXISTS _mycobank_columns")

    cols_sql = ", ".join(f"{quote_ident(column)} TEXT" for column in columns)
    cur.execute(f"CREATE TABLE IF NOT EXISTS {quote_ident(table_name)} ({cols_sql})")
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS _mycobank_import_meta (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS _mycobank_columns (
            position INTEGER PRIMARY KEY,
            column_name TEXT NOT NULL,
            original_header TEXT
        )
        """
    )
    cur.executemany(
        """
        INSERT OR REPLACE INTO _mycobank_columns
            (position, column_name, original_header)
        VALUES (?, ?, ?)
        """,
        [(idx + 1, column, original_headers[idx]) for idx, column in enumerate(columns)],
    )


def insert_metadata(
    conn: sqlite3.Connection,
    *,
    source_zip: Path,
    source_url: str | None,
    zip_member: str,
    excel_path: Path,
    sheet_name: str,
    table_name: str,
    row_count: int,
) -> None:
    metadata = {
        "imported_at": datetime.now(timezone.utc).isoformat(),
        "source_url": source_url or "",
        "source_zip": str(source_zip),
        "source_zip_sha256": sha256_file(source_zip),
        "source_zip_size_bytes": str(source_zip.stat().st_size),
        "zip_member": zip_member,
        "excel_file": str(excel_path),
        "excel_sha256": sha256_file(excel_path),
        "excel_size_bytes": str(excel_path.stat().st_size),
        "sheet_name": sheet_name,
        "table_name": table_name,
        "row_count": str(row_count),
    }
    conn.executemany(
        """
        INSERT OR REPLACE INTO _mycobank_import_meta (key, value)
        VALUES (?, ?)
        """,
        metadata.items(),
    )


def create_indexes(
    conn: sqlite3.Connection,
    *,
    table_name: str,
    index_columns: list[str],
) -> None:
    cur = conn.cursor()
    available_columns = {
        row[1].lower(): row[1]
        for row in cur.execute(f"PRAGMA table_info({quote_ident(table_name)})")
    }

    for requested_column in index_columns:
        normalized = requested_column.lower()
        if normalized not in available_columns:
            print(
                f"[WARN] No existe la columna para indexar: {requested_column}",
                file=sys.stderr,
            )
            continue
        column = available_columns[normalized]
        index_name = f"idx_{table_name}_{column}"
        cur.execute(
            f"""
            CREATE INDEX IF NOT EXISTS {quote_ident(index_name)}
            ON {quote_ident(table_name)} ({quote_ident(column)})
            """
        )


def import_excel_to_sqlite(
    *,
    excel_path: Path,
    db_path: Path,
    zip_path: Path,
    zip_member: str,
    source_url: str | None,
    table_name: str = DEFAULT_TABLE,
    sheet_name: str = DEFAULT_SHEET,
    replace: bool = True,
    batch_size: int = 10_000,
    index_columns: list[str] | None = None,
) -> int:
    """Import the selected Excel workbook into SQLite and return row count."""
    db_path.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(filename=excel_path, read_only=True, data_only=True)

    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"No existe la hoja {sheet_name!r}. "
                f"Hojas encontradas: {', '.join(wb.sheetnames)}"
            )

        extras = [
            name
            for name in wb.sheetnames
            if name != sheet_name and worksheet_has_data(wb[name])
        ]
        if extras:
            print(
                "[WARN] Se han encontrado hojas adicionales no vacías: "
                + ", ".join(extras),
                file=sys.stderr,
            )

        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        raw_header = get_header_row(rows)
        original_headers = ["" if value is None else str(value).strip() for value in raw_header]
        used_columns: set[str] = set()
        columns = [
            clean_column_name(value, idx + 1, used_columns)
            for idx, value in enumerate(raw_header)
        ]

        conn = sqlite3.connect(db_path)
        try:
            conn.execute("PRAGMA journal_mode = WAL")
            conn.execute("PRAGMA synchronous = OFF")
            conn.execute("PRAGMA temp_store = MEMORY")
            prepare_database(
                conn,
                table_name=table_name,
                columns=columns,
                original_headers=original_headers,
                replace=replace,
            )

            placeholders = ", ".join("?" for _ in columns)
            insert_sql = (
                f"INSERT INTO {quote_ident(table_name)} "
                f"({', '.join(quote_ident(column) for column in columns)}) "
                f"VALUES ({placeholders})"
            )
            row_count = 0
            batch: list[tuple[object, ...]] = []
            for row in rows:
                values = list(row)
                if len(values) < len(columns):
                    values.extend([None] * (len(columns) - len(values)))
                elif len(values) > len(columns):
                    values = values[: len(columns)]
                if not any(value is not None for value in values):
                    continue

                batch.append(tuple(sqlite_value(value) for value in values))
                if len(batch) >= batch_size:
                    conn.executemany(insert_sql, batch)
                    row_count += len(batch)
                    batch.clear()

            if batch:
                conn.executemany(insert_sql, batch)
                row_count += len(batch)

            create_indexes(
                conn,
                table_name=table_name,
                index_columns=index_columns or [],
            )
            insert_metadata(
                conn,
                source_zip=zip_path,
                source_url=source_url,
                zip_member=zip_member,
                excel_path=excel_path,
                sheet_name=sheet_name,
                table_name=table_name,
                row_count=row_count,
            )
            conn.commit()
            return row_count
        finally:
            conn.close()
    finally:
        wb.close()


def build_database(
    *,
    db_path: Path,
    zip_path: Path | None = None,
    url: str = DEFAULT_MYCOBANK_URL,
    cache_dir: Path | None = None,
    table_name: str = DEFAULT_TABLE,
    sheet_name: str = DEFAULT_SHEET,
    replace: bool = True,
    refresh: bool = False,
    batch_size: int = 10_000,
    index_columns: list[str] | None = None,
) -> Path:
    """Build a MycoBank SQLite database from a URL or a local ZIP."""
    resolved_cache = cache_dir or (Path.home() / ".cache" / DEFAULT_CACHE_DIR_NAME)
    resolved_zip = resolve_zip_path(
        zip_path=zip_path,
        url=url,
        cache_dir=resolved_cache,
        refresh=refresh,
    )
    source_url = None if zip_path is not None else url

    with tempfile.TemporaryDirectory(prefix="pycobank_") as tmp:
        excel_path, zip_member = extract_mblist_excel(resolved_zip, Path(tmp))
        rows = import_excel_to_sqlite(
            excel_path=excel_path,
            db_path=db_path,
            zip_path=resolved_zip,
            zip_member=zip_member,
            source_url=source_url,
            table_name=table_name,
            sheet_name=sheet_name,
            replace=replace,
            batch_size=batch_size,
            index_columns=index_columns or [
                "taxon_name",
                "current_name",
                "mycobank",
                "current_mycobank",
                "rank",
            ],
        )

    print(f"SQLite creado: {db_path}")
    print(f"Tabla principal: {table_name}")
    print(f"Filas importadas: {rows}")
    return db_path
